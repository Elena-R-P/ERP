from .models import Driver, Truck, Trailer, Load, Payroll


def import_database():
    # pip install openpyxl
    from openpyxl import load_workbook

    wb = load_workbook(filename="/src/staticfiles/2025_SafetyDepartment.xlsx")
    ws = wb["2025_SafetyDepartment.xlsx"]
    # for row in ws.iter_rows(min_row=2, value_only=True):
